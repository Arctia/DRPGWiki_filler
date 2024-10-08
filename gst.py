
import openpyxl, json, sys, re, os

from plibs.config import Config
from openpyxl.styles import PatternFill
from globals import *
from xlsx_fetcher import download_excell

jconf = Config("./plibs/")

wpath = os.path.join("translation_sheet", "Disgaea RPG Translations of Characters.xlsx")
spath = os.path.join("translation_sheet", "result.xlsx")

if "--refresh" in sys.argv:
	if not download_excell():
		print("[ERROR ]: Cannot download xlsx file.. Using old one")
else:
	print("[INFO  ]: Using old xlsx")

work_to_do = None
if "--all" in sys.argv:
	work_to_do = "all"
elif "--write" in sys.argv:
	work_to_do = "write"
elif "--update" in sys.argv:
	work_to_do = "update"
elif "--buffed" in sys.argv:
	work_to_do = "buffed"

try:
	workbook = openpyxl.load_workbook(wpath)
except:
	print("cannot load excell file")
#ws["A1"].fill = PatternFill("solid", start_color="FFA500")

# Row | Col | Text

RowIDS = {"Character ID": 1,
		  "Character Name": 2,
		  "Bio": 3,
		  "Title 1": 4,
		  "Title 2": 5,
		  "Title 3": 6,
		  "Title 4": 7,
		  "Title 5": 8,
		  "Title 6": 9,
		  "Skill Name 1": 11,
		  "Skill Desc 1": 12,
		  "Skill Effect 1": 13,
		  "Skill Name 2": 14,
		  "Skill Desc 2": 15,
		  "Skill Effect 2": 16,
		  "Skill Name 3": 17,
		  "Skill Desc 3": 18,
		  "Skill Effect 3": 19,
		  "Skill Name 4": 20,
		  "Skill Desc 4": 21,
		  "Skill Effect 4": 22,
		  "Evility Name 1": 24,
		  "Evility Desc 1": 25,
		  "Evility Name 2": 26,
		  "Evility Desc 2": 27,
		  "Evility Name 3": 28,
		  "Evility Desc 3": 29,
		  "Evility Name 4": 30,
		  "Evility Desc 4": 31,
		  }

ids_to_check = [13,16,19,22,25,27,29,31]

offset_x = 2
cd = []
UNIQUE_HUMAN = []
UNIQUE_HUMAN_2 = []
UNIQUE_MONSTER = []
UNIQUE_UNPLAYABLE = []
HUMAN = []
MONSTER = []

offset_y = 33

jish = {'Unique Human': UNIQUE_HUMAN,
		'Unique Human 2': UNIQUE_HUMAN_2,
		'Unique Monster': UNIQUE_MONSTER, 
		'Unique Unplayable': UNIQUE_UNPLAYABLE,
		"Human Generic": HUMAN, 
		"Monster Generic": MONSTER}

def replace_wrong_evility():
	for c in Charas:
		if c["id"] != 127: continue
		c["additional_m_leader_skill_id_sub_3"] = 1278
		c["m_leader_skill_id_sub_3"] = 1274
		break
	for e in LeaderSkills:
		if e["id"] == 1274: e["id"] = 1278
		elif e["id"] == 1278: e["id"] = 1274

replace_wrong_evility()

def main():
	for c in Charas:
		if c["id"] >= 50000: continue
		l = {}
		l = Infos(c, l)
		l = Skills(c, l)
		l = Evilities(c, l)
		Fill_Arrays(c,l)

	jish = {'Unique Human': UNIQUE_HUMAN,
		'Unique Human 2': UNIQUE_HUMAN_2,
		'Unique Monster': UNIQUE_MONSTER, 
		'Unique Unplayable': UNIQUE_UNPLAYABLE,
		"Human Generic": HUMAN, 
		"Monster Generic": MONSTER}
	
	AddExcell(jish)

def modify():
	for c in Charas:
		if c["id"] >= 50000: continue
		l = {}
		l = Infos(c, l)
		l = Skills(c, l)
		l = Evilities(c, l)
		Fill_Arrays(c,l)

	jish = {'Unique Human': UNIQUE_HUMAN,
		'Unique Human 2': UNIQUE_HUMAN_2,
		'Unique Monster': UNIQUE_MONSTER, 
		'Unique Unplayable': UNIQUE_UNPLAYABLE,
		"Human Generic": HUMAN, 
		"Monster Generic": MONSTER}
	
	ModifyExcell(jish)

def blank_cells():
	for c in Charas:
		if c["id"] >= 50000: continue
		l = {}
		l = Infos(c, l)
		l = Skills(c, l)
		l = Evilities(c, l)
		Fill_Arrays(c,l)

	jish = {'Unique Human': UNIQUE_HUMAN,
		'Unique Human 2': UNIQUE_HUMAN_2,
		'Unique Monster': UNIQUE_MONSTER, 
		'Unique Unplayable': UNIQUE_UNPLAYABLE,
		"Human Generic": HUMAN, 
		"Monster Generic": MONSTER}
	
	BlankExcell(jish)

def read():

	for c in Charas:
		l = {}
		l = Infos(c, l)
		l = Skills(c, l)
		l = Evilities(c, l)
		Fill_Arrays(c,l)

	jish = {'Unique Human': UNIQUE_HUMAN,
		'Unique Human 2': UNIQUE_HUMAN_2,
		'Unique Monster': UNIQUE_MONSTER, 
		'Unique Unplayable': UNIQUE_UNPLAYABLE,
		"Human Generic": HUMAN, 
		"Monster Generic": MONSTER}

	cids = ReadExcell(jish)

	start = f"""<includeonly>{{{{#switch: {{{{{{1}}}}}}\n"""
	names, bios, evilities, evilities_effect, titles, skills, skills_desc, skills_eff = start, start, start, start, start, start, start, start
	for c in Charas:
		for i in cids:
			try:
				if cids[i]["Character ID"] == c["id"]:
					ci = cids[i]
					names += WriteNames(c, ci)
					bios += WriteBios(c, ci)
					evilities += WriteEvilities(c, ci)
					evilities_effect += WriteEvilitiesEffect(c, ci)
					titles += WriteTitles(c, ci)
					skills += WriteSkills(c, ci)
					skills_desc += WriteSkillsDescription(c, ci)
					skills_eff += WriteSkillsEffect(c, ci)
			except:
				pass

	end = f"""|}}}}</includeonly><noinclude>\nThis page serves for translations, if you want to help with that please contact Arctia before modify this page directly. Translations are made by [[User:DjinnandTonic|Djinn]], [[User:XYZexal|XYZexal]], Moot, [[User:DurianNoodle|DurianNoodle]] and google translate.\n[[Category:Enum]][[Category:TranslationTables]]\n</noinclude>"""
	print("\nUploading sheets...")
	with open("./translation_sheet/txt/names.txt", "w", encoding="utf-8") as f:
		f.write(names + end)
		page = wiki.pages["Template:Enum/JP/CharaName"]
		content = names + end
		Upload(page, content, True)
	with open("./translation_sheet/txt/bios.txt", "w", encoding="utf-8") as f:
		f.write(bios + end)
		page = wiki.pages["Template:Enum/JP/Bios"]
		content = bios + end
		Upload(page, content, True)
	with open("./translation_sheet/txt/evilities.txt", "w", encoding="utf-8") as f:
		f.write(evilities + end)
		page = wiki.pages["Template:Enum/JP/LeaderSkills"]
		content = evilities + end
		Upload(page, content, True)
	with open("./translation_sheet/txt/titles.txt", "w", encoding="utf-8") as f:
		f.write(titles + end)
		page = wiki.pages["Template:Enum/JP/Titles"]
		content = titles + end
		Upload(page, content, True)
	with open("./translation_sheet/txt/skills.txt", "w", encoding="utf-8") as f:
		f.write(skills + end)
		page = wiki.pages["Template:Enum/JP/Skills"]
		content = skills + end
		Upload(page, content, True)
	with open("./translation_sheet/txt/skills_description.txt", "w", encoding="utf-8") as f:
		f.write(skills_desc + end)
		page = wiki.pages["Template:Enum/JP/SkillsDescription"]
		content = skills_desc + end
		Upload(page, content, True)
	with open("./translation_sheet/txt/evilities_effect.txt", "w", encoding="utf-8") as f:
		f.write(evilities_effect + end)
		page = wiki.pages["Template:Enum/JP/LeaderSkills/Effect"]
		content = evilities_effect + end
		Upload(page, content, True)
	with open("./translation_sheet/txt/skills_effect.txt", "w", encoding="utf-8") as f:
		f.write(skills_eff + end)
		page = wiki.pages["Template:Enum/JP/SkillsEffect"]
		content = skills_eff + end
		Upload(page, content, True)
	print("Sheets uploaded!!!")

def Fill_Arrays(c, l):
	if c["unique_flg"] == True:
		if c["m_leader_skill_id"] == 0 or c['id']>40000: UNIQUE_UNPLAYABLE.append(l)
		elif c["character_type"] == 1:
			if c["id"] < 200: UNIQUE_HUMAN.append(l)
			else: UNIQUE_HUMAN_2.append(l)
		elif c["character_type"] == 2: UNIQUE_MONSTER.append(l)
	elif c["character_type"] == 1: HUMAN.append(l)
	elif c["character_type"] == 2: MONSTER.append(l)

def WriteNames(c, ci):
	return f'| {c["id"]} = {ci["Character Name"]}\n'

def WriteBios(c, ci):
	return f'| {c["id"]} = {ci["Bio"]}\n'

def WriteEvilities(c, ci):
	cid = c['id']
	l_ids = [c["m_leader_skill_id"], c["m_leader_skill_id_sub_1"], c["m_leader_skill_id_sub_2"], c["m_leader_skill_id_sub_3"]]
	string = f""
	count = 1
	for lid in l_ids:
		for e in LeaderSkills:
			if e["id"] != lid: continue
			
			string += f"""| {lid} = {ci[f"Evility Name {count}"]}\n"""
			count += 1
			break
	return string

def EvilityEffect(e, effect):
	_to_replace = ["# PER #", "# PER#", "#PER #", "#PER#"]
	if e["id"] in [542, 1672, 100262, 100161]: # Drop Rate Increase exception
		val = (e["effect_value_min"]/100)+1
		for r in _to_replace: effect = effect.replace(r, str(val))
		return effect
	for r in _to_replace: effect = effect.replace(r, str(e["effect_value_min"]))
	return effect

def SettingPerString(effect):
	_to_replace = ["# PER #", "# PER#", "#PER #"]
	_to_replace2 = ["# PER2 #", "# PER2#", "#PER2 #"]
	for r in _to_replace: effect = effect.replace(r, "#PER#")
	for r in _to_replace2: effect = effect.replace(r, "#PER2#")
	return effect

def EvilityEffectSpecialCase(e, effect, add=0):
	par = "#PER#" if add == 0 else "#PER2#"
	if e["id"] in [542, 1672, 100262, 100161]:
		val = (e["effect_value_min"]/100)+1
		effect = effect.replace(par, str(val))
	return effect

def EvilityEffectReplacePer(e, effect, add=0):
	par = "#PER#" if add == 0 else "#PER2#"
	effect = effect.replace(par, str(e["effect_value_min"]))
	return effect

def EvilityEffectAdditional(e, effect, ae):
	effect = SettingPerString(effect)
	effect = EvilityEffectSpecialCase(e, effect, 0)
	effect = EvilityEffectSpecialCase(ae, effect, 1)
	effect = EvilityEffectReplacePer(e, effect, 0)
	effect = EvilityEffectReplacePer(ae, effect, 1)
	return effect

def WriteEvilitiesEffect(c, ci):
	l_ids = [c["m_leader_skill_id"], c["m_leader_skill_id_sub_1"], c["m_leader_skill_id_sub_2"], c["m_leader_skill_id_sub_3"]]
	olids = [c["additional_m_leader_skill_id"], c["additional_m_leader_skill_id_sub_1"], c["additional_m_leader_skill_id_sub_2"], c["additional_m_leader_skill_id_sub_3"]]
	string = f""
	count = 1

	for lids in l_ids:
		for e in LeaderSkills:
			if e["id"] != lids: continue
			if olids[count-1] != 0:
				#print(l_ids + olids) #Printing Double IDs
				for la in LeaderSkills:
					if la["id"] != olids[count-1]: continue
					effect = EvilityEffectAdditional(e, ci[f"Evility Desc {count}"], la)
					string += f"""| {lids} = {effect}\n"""
					break
				break
			else:
				effect = EvilityEffect(e, ci[f"Evility Desc {count}"])
				string += f"""| {lids} = {effect}\n"""
				break
		count += 1

	return string

def WriteTitles(c, ci):
	count = 1
	string = ""
	not_found = True
	for cl in ClassNames:
		if cl["m_character_id"] == c["id"]:
			string += f"""| {c['id']}{count} = {ci[f"Title {count}"]}\n"""
			count +=1
			not_found = False
			if count == 7: break
	return string

def WriteSkills(c, ci):
	skills_ids = []
	for s in CharaCommands:
		if s["m_character_id"] == c["id"]:
			skills_ids.append(s["m_command_id"])

	for s in CharaRetrofits:
		if s["m_character_id"] == c["id"]:
			if s["retrofit_type"] == 1:
				skills_ids.append(s["retrofit_value"])
				break

	string = ""
	count = 1
	for sid in skills_ids:
		for s in Commands:
			if s["id"] == sid:
				string += f"""| {s['id']} = {ci[f"Skill Name {count}"]}\n"""
				count += 1
				break
	return string

def WriteSkillsDescription(c, ci):
	skills_ids = []
	for s in CharaCommands:
		if s["m_character_id"] == c["id"]:
			skills_ids.append(s["m_command_id"])

	for s in CharaRetrofits:
		if s["m_character_id"] == c["id"]:
			if s["retrofit_type"] == 1:
				skills_ids.append(s["retrofit_value"])
				break

	string = ""
	count = 1
	for sid in skills_ids:
		for s in Commands:
			if s["id"] == sid:
				string += f"""| {s['id']} = {ci[f"Skill Desc {count}"]}\n"""
				count += 1
				break
	return string

def WriteSkillsEffect(c, ci):
	skills_ids = []
	for s in CharaCommands:
		if s["m_character_id"] == c["id"]:
			skills_ids.append(s["m_command_id"])

	for s in CharaRetrofits:
		if s["m_character_id"] == c["id"]:
			if s["retrofit_type"] == 1:
				skills_ids.append(s["retrofit_value"])
				break

	string = ""
	count = 1
	for sid in skills_ids:
		for s in Commands:
			if s["id"] == sid:
				effect = ci[f"Skill Effect {count}"]

				for e in range(1,10):
					effect = effect.replace("# ", "#").replace("%s #" % e, "%s#" % e)
					if ("#PER%s#" % e) in effect:
						if s["effect_values_min"][e-1] != s["effect_values_max_50"][e-1]:
							value = "%s>%s" % (s["effect_values_min"][e-1], s["effect_values_max_50"][e-1])
							value = value.replace("%", "", 1)
						else:
							value = "%s" % s["effect_values_min"][e-1]
						effect = effect.replace("#PER%s#" % e, value)
				for z in ["+", "-"]:
					if z in effect:
						effect = effect.replace(z, " " +z).replace("%", "% ").replace("  ", " ").replace(" ,", ", ").replace("  ", " ")

				string += f"""| {s['id']} = {effect}\n"""
				count +=1
				break
	return string

def Infos(c, l, res = {}):
	l["Character ID"] = c["id"]
	l["Character Name"] = c["name"]
	l["Bio"] = c["description"]
	l['book_appear_at'] = c["book_appear_at"]
	count = 1
	for cl in ClassNames:
		if cl["m_character_id"] == c["id"]:
			l[f"Title {count}"] = cl["class_name"]
			count += 1

	return l

def Skills(c, l):
	skills_ids = []
	for s in CharaCommands:
		if s["m_character_id"] == c["id"]:
			skills_ids.append(s["m_command_id"])

	for s in CharaRetrofits:
		if s["m_character_id"] == c["id"]:
			if s["retrofit_type"] == 1:
				skills_ids.append(s["retrofit_value"])
				break

	count = 1
	for sid in skills_ids:
		for s in Commands:
			if s["id"] == sid:
				l[f"Skill Name {count}"] = s["name"]
				l[f"Skill Desc {count}"] = s["description"]
				l[f"Skill Effect {count}"] = s["description_effect"]
				count += 1
				break

	return l

def Evilities(c, l):
	l_ids = [c["m_leader_skill_id"], c["m_leader_skill_id_sub_1"], c["m_leader_skill_id_sub_2"], c["m_leader_skill_id_sub_3"]]
	olids = [c["additional_m_leader_skill_id"], c["additional_m_leader_skill_id_sub_1"], c["additional_m_leader_skill_id_sub_2"], c["additional_m_leader_skill_id_sub_3"]]

	if l_ids[3] > olids[3] and olids[3] != 0:
		# print("first leaderskill is after")
		swap = l_ids[3]
		l_ids[3] = olids[3]
		olids[3] = swap

	count = 1
	for lid in l_ids:
		for e in LeaderSkills:
			if e["id"] != lid: continue
			l[f"Evility Name {count}"] = e["name"]
			if olids[count-1] != 0:
				for ae in LeaderSkills:
					if ae["id"] != olids[count-1]: continue
					des = ae["description"].replace("#PER#", "#PER2#")[1:]
					l[f"Evility Desc {count}"] = f"{e['description']}, {des}"
					# if c["id"] == 200: 
					# 	# print(l[f"Evility Desc {count}"])
					# 	pass
					break 
			else:
				l[f"Evility Desc {count}"] = e["description"]
			count += 1
			break

	return l

def spells():
	if JP:
		col = offset_x
		worksheet = workbook["Spells"]
		for c in Commands:
			if c["id"] < 200000 or c["id"] > 299999: continue
			worksheet.cell(1, col, "Japanese")
			worksheet.cell(1, col+1, "English")
			counter = 2
			for e in ["id", "name", "description", "description_effect"]:
				worksheet.cell(counter, col, c[e])
				counter += 1
			col += 3
	else:
		print("Using Global Spells")
		col = offset_x + 1
		worksheet = workbook["Spells"]
		for c in Commands:
			if c["id"] < 200000 or c["id"] > 299999: continue
			counter = 2
			for e in ["id", "name", "description", "description_effect"]:
				worksheet.cell(counter, col, c[e])
				counter += 1
			col += 3
	workbook.save(spath)

def WriteExcell(jish):
	if JP:
		for key in jish:
			col = offset_x
			worksheet = workbook[key]
			for c in jish[key]:
				worksheet.cell(1, col, "Japanese")
				worksheet.cell(1, col+1, "English")
				for i in c:
					worksheet.cell(RowIDS[i]+1, col, c[i])
				col += 3
	else:
		for key in jish:
			col = offset_x
			worksheet = workbook[key]
			for c in jish[key]:
				for cl in range(2,worksheet.max_column+1, 3):
					if worksheet.cell(RowIDS["Character ID"]+1, cl).value == c["Character ID"]:
						for i in c:
							worksheet.cell(RowIDS[i]+1, cl+1, c[i])

	
	workbook.save(spath)

def AddExcell(jish):
	pre_cell = 0
	if JP:
		for key in jish:
			if key == 'Unique Unplayable': continue
			col = offset_x
			worksheet = workbook[key]
			for c in jish[key]:
				found = False
				last_cell = 0
				for cl in range(2, worksheet.max_column+3, 3):
					cell_value = worksheet.cell(RowIDS["Character ID"]+1, cl).value
					if cell_value == c["Character ID"]: break
					if cell_value == None or cell_value == "None": last_cell = cl;break
					if cell_value > c["Character ID"]:
						found = True
						# move the columns
						for i in range(3):
							worksheet.insert_cols(cl)
						for f in c:
							if f != "book_appear_at": worksheet.cell(RowIDS[f]+1, cl, c[f])
						print("[INFO	]: Added character %s" % c["Character ID"])
						jconf.add_new_chara(c)
						break
				if not found and last_cell > 0:
					cl = last_cell
					for i in range(3):
						worksheet.insert_cols(cl)
					for f in c:
						if f != "book_appear_at": worksheet.cell(RowIDS[f]+1, cl, c[f])
					print("[INFO	]: Added character %s" % c["Character ID"])
					jconf.add_new_chara(c)

	workbook.save(spath)


def push_down_char(ws, cl:int) -> None:
	row = offset_y
	while ws.cell(row + 2, cl).value != None: 
		row += offset_y
	
	for field in RowIDS.values():
		field += 1
		japanese_value = ws.cell(field, cl).value
		translation_value = ws.cell(field, cl + 1).value
		ws.cell(row + field, cl, japanese_value)
		ws.cell(row + field, cl + 1, translation_value)


modded_chars = []
def ModifyExcell(jish):
	if JP:
		for key in jish:
			worksheet = workbook[key]
			for c in jish[key]:
				for cl in range(2, worksheet.max_column+3, 3):
					cell_value = worksheet.cell(RowIDS["Character ID"]+1, cl).value
					#if cell_value == c["Character ID"]: break
					if cell_value == c["Character ID"]:
						for f in c:
							if f == "book_appear_at": continue
							value = worksheet.cell(RowIDS[f]+1, cl).value
							if value != c[f] and RowIDS[f] in ids_to_check:
								if not c["Character ID"] in modded_chars and c["Character ID"] == 242 and f == "Skill Effect 3":
									continue
								if not c["Character ID"] in modded_chars:
									print("[INFO	]: Modified character %s" % c["Character ID"])
									push_down_char(worksheet, cl)
									modded_chars.append(c["Character ID"])
								# fill row with new info and color it
								worksheet.cell(RowIDS[f]+1, cl, c[f])
								worksheet.cell(RowIDS[f]+1, cl).fill = PatternFill("solid", start_color="20EE20")
								worksheet.cell(RowIDS[f]+1, cl + 1, "")
								jconf.add_mod_chara(c)
						break

	workbook.save(spath)

def BlankExcell(jish):
	pre_cell = 0
	if JP:
		for key in jish:
			col = offset_x
			worksheet = workbook[key]
			for c in jish[key]:
				found = False
				last_cell = 0
				for cl in range(2, worksheet.max_column+3, 3):
					cell_value = worksheet.cell(RowIDS["Character ID"]+1, cl).value
					#if cell_value == c["Character ID"]: break
					if cell_value == None or cell_value == "None": last_cell = cl;break
					if cell_value == c["Character ID"]:
						found = True
						for f in c:
							if f == "book_appear_at": continue
							if RowIDS[f]+1 == 2: continue
							value = worksheet.cell(RowIDS[f]+1, cl).value
							t_value = worksheet.cell(RowIDS[f]+1, cl+1).value
							if not value in [None, "None", "-"]:
								if t_value in [None, "None"]:
									worksheet.cell(RowIDS[f]+1, cl).fill = PatternFill("solid", start_color="bf819e")
									print("[INFO	]: Modified character %s" % c["Character ID"])
							#if value != c[f]:
							#	worksheet.cell(RowIDS[f]+1, cl, c[f])
							#	worksheet.cell(RowIDS[f]+1, cl).fill = PatternFill("solid", start_color="F6CDFF")
							#	print("[INFO	]: Modified character %s" % c["Character ID"])
						break

	workbook.save(spath)

class Record():
	cids_check = {}
	cids_record = {}

	def load_json(self):
		with open('./translation_sheet/save/check.json') as f:
			self.cids_check = json.load(f)
		with open('./translation_sheet/save/translated.json') as f:
			self.cids_record = json.load(f)

	def save_json(self):
		with open('./translation_sheet/save/check.json', 'w') as json_file:
			json.dump(self.cids_check, json_file, indent=4)
		with open('./translation_sheet/save/translated.json', 'w') as json_file:
			json.dump(self.cids_record, json_file, indent=4)

	def WriteRecord(self, c, key, val):
		#print(cids_check)
		if not str(c["Character ID"]) in self.cids_check: self.cids_check[str(c["Character ID"])] = {}
		self.cids_check[str(c["Character ID"])][key] = val

	def WriteTranslation(self,c, key, val):
		if not str(c["Character ID"]) in self.cids_record: self.cids_record[str(c["Character ID"])] = {}
		self.cids_record[str(c["Character ID"])][key] = val

	def CheckRecord(self,c, key, val):
		if str(c["Character ID"]) in self.cids_check:
			if key in self.cids_check[str(c["Character ID"])]:
				if self.cids_check[str(c["Character ID"])][key] == val: return True
		return False

	def ReturnTranslation(self,c, key):
		if str(c["Character ID"]) in self.cids_record:
			#print(cids_record[str(c["Character ID"])][key])
			return self.cids_record[str(c["Character ID"])][key]
		return None

class AI_Translations():
	ai_translated = None

	def load_json(self):
		with open('./translation_sheet/AI/ai_translated.json') as f:
			self.ai_translated = json.load(f)

	def get_chara(self, id:int):
		for c in self.ai_translated:
			if int(c["Character ID"]) == id:
				return c
		return None

	def check_value(self, c, key:str):
		chara = self.get_chara(c["Character ID"])
		if not chara: return None
		return chara[key]

def replacenth(string, sub, wanted, n):
    where = [m.start() for m in re.finditer(sub, string)][n-1]
    before = string[:where]
    after = string[where:]
    after = after.replace(sub, wanted, 1)
    newString = before + after
    return newString

RECORD = Record()
AI_TRANSLATIONS = AI_Translations()

def ReadExcell(jish):
	cids = {}
	RECORD.load_json()
	AI_TRANSLATIONS.load_json()
	#print(RECORD.cids_check)
	for key in jish:
		col = offset_x
		worksheet = workbook[key]
		for c in jish[key]:
			_hash = {}
			for cl in range(2,worksheet.max_column+1, 3):
				if worksheet.cell(RowIDS["Character ID"]+1, cl).value == c["Character ID"]:
					for key in RowIDS:
						if "Character ID" in key: continue # or "Skill Effect" in key: continue
						value = worksheet.cell(RowIDS[key]+1, cl+1).value
						if value == None or value == "None":
							val = worksheet.cell(RowIDS[key]+1, cl).value
							if not val in ["None", None, "", " ", "  "]:
								# check AI translated stuff
								tmp_val = AI_TRANSLATIONS.check_value(c, key)
								if tmp_val != None:
									value = tmp_val
								elif RECORD.CheckRecord(c, key, val) and not c["Character ID"] in [200374]:
									# IF Japanese string didn't change load the already translated text
									value = RECORD.ReturnTranslation(c, key)
								else:
									RECORD.WriteRecord(c, key, val)
									# print("Japanese: " + val)
									value = translate(val)
									if value.count("#PER#") > 1 and "Evility Desc" in key:
										value = replacenth(value, "#PER#", "#PER2#", 2)
									# print("English: " + value)
									RECORD.WriteTranslation(c, key, value)
							else:
								RECORD.WriteRecord(c, key, val)
						_hash[key] = value
					print(f'Chara id: {c["Character ID"]}')

					_hash["Character ID"] = c["Character ID"] 
					break
			if _hash != {}: cids[c["Character ID"]] = _hash
	RECORD.save_json()
	return cids

if __name__ == "__main__":
	if work_to_do == None:
		_r_ = input("""write:
 w to write result.xlxs
 r to read values and push changes
 s to write spells
 m to check cells buffed or changed with green
 p to check cells untranslated with violet
what do you want to do? """)
		if _r_ == "w": main()
		elif _r_ == "r": read()
		elif _r_ == "s": spells()
		elif _r_ == "m": modify()
		elif _r_ == "p": blank_cells()
		elif _r_ == "e": exit()
		exit(0)

	if work_to_do == "all":
		main()
		workbook = openpyxl.load_workbook(spath)
		modify()
	elif work_to_do == "write": main()
	elif work_to_do == "buffed": modify()
	elif work_to_do == "update": read() 